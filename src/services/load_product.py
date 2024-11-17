import re

from aiogram.types import Message
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.database.db_functions import DB


def load_iphone_from_xlsx(path_file: str, message: Message):
    wb: Workbook = load_workbook(path_file)
    ws: Worksheet = wb.active

    max_row = ws.max_row + 1

    for row_number in range(2, max_row):
        try:
            product_type_name = str(ws.cell(row=row_number, column=1).value).title()
            manufacturer_name = str(ws.cell(row=row_number, column=2).value).title()
            subtype_name = str(ws.cell(row=row_number, column=3).value).title()
            product_name = ws.cell(row=row_number, column=4).value
            price = int(ws.cell(row=row_number, column=5).value)
            quantity = int(ws.cell(row=row_number, column=6).value)

            product_type_id = DB.create_product_type_if_not_exists(product_type_name)
            manufacturer_id = DB.create_manufacturer_company_if_not_exists(manufacturer_name)
            subtype_id = DB.create_subtype_if_not_exists(subtype_name, product_type_id, manufacturer_id)

            DB.save_product(product_name, price, subtype_id, quantity)
        except Exception as e:
            message.answer(f"Что-то пошло не так в строке {row_number}. Она не будет добавлена в бд.\n"
                           f"Error: {e}")


def parse_to_int(text: str) -> int:
    int_lst = re.findall(r"\d+", text)
    number = "".join(int_lst)
    return int(number)
